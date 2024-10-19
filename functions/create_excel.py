import os
from utils.models import Races
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Alignment, PatternFill, Border, Side


async def create_excel(dict_list: list[Races], file_name: str, sheet_name: str = 'Статистика за выбранный период'):
    if not os.path.exists('excel_files'):
        os.mkdir('excel_files')

    filepath = os.path.join('excel_files', f'{file_name}.xlsx')

    # Создаем новую книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Записываем данные из списка словарей в Excel
    if dict_list:
        header = ['ID рейса', 'ID пользователя', 'Имя водителя', 'Номер ТС', 'Начало рейса', 'Количество точек', '1',
                  '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15']
        ws.append(header)  # Записываем заголовки

        for row in dict_list:
            ws.append([row.id, row.user_id, row.fullname, row.number_car, row.datetime_start, row.points,
                       *row.points_time.split(';')])

    # Настраиваем стили для красивого вида
    header_style = NamedStyle(name='header')
    header_style.font = Font(bold=True, color='FFFFFF')
    header_style.alignment = Alignment(horizontal='center', vertical='center')
    header_style.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    border_style = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )
    header_style.border = border_style

    cell_style = NamedStyle(name='cell')
    cell_style.alignment = Alignment(horizontal='left', vertical='center')
    cell_style.border = border_style

    for cell in ws[1]:  # Применяем стиль к заголовкам
        cell.style = header_style

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.style = cell_style

    # Автоматическое изменение ширины столбцов
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Сохраняем файл
    wb.save(filepath)
    return filepath
